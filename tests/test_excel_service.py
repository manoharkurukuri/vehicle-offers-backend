from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from app.schemas.llm import VehicleIncentiveLLM
from app.services.excel_service import ExcelService


def test_file_name_and_workbook(tmp_path):
    service = ExcelService(
        storage_dir=tmp_path,
        timezone_name="America/Los_Angeles",
    )

    created_at = datetime(2026, 8, 11, 9, 0, tzinfo=ZoneInfo("America/Los_Angeles"))
    assert (
        service.build_file_name("Norm Reeves Honda Irvine", created_at)
        == "norm_reeves_honda_irvine_august_11_tuesday.xlsx"
    )

    incentive = VehicleIncentiveLLM.model_validate(
        {
            "offer_priority": "Vehicle #1",
            "offer_type": "Lease Offer",
            "vehicle_type": "New",
            "year": 2026,
            "make": "Honda",
            "model": "Accord Sedan",
            "trim": "LX",
            "drive_train": "FWD",
            "lowest_monthly_payment": 229,
            "lease_term_months": 36,
            "down_payment_or_due_at_signing": "Due at Signing",
            "total_due_at_signing": 3998,
            "annual_mileage": 10000,
        }
    )

    file_name, path = service.create_workbook(
        "Norm Reeves Honda Irvine",
        [incentive],
        source_url="https://www.normreeveshondairvine.com/vehicle-specials/",
    )

    assert file_name.endswith(".xlsx")
    workbook = load_workbook(path)
    sheet = workbook["Monthly Vehicle Incentives"]

    assert sheet.max_column == 26
    assert sheet["A2"].value == "Vehicle #1"
    assert sheet["B2"].value == "Lease Offer"
    assert sheet["M2"].value == 229
    assert sheet["O2"].value == "Due at Signing"
    assert sheet["Q2"].value == 3998
    assert sheet["A1"].comment is not None
    assert "https://www.normreeveshondairvine.com/vehicle-specials/" in sheet["A1"].comment.text
    assert len(sheet.data_validations.dataValidation) == 3
