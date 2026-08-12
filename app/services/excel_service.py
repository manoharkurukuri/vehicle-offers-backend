import os
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.config import settings
from app.core.constants import (
    EXCEL_FIELD_ORDER,
    EXCEL_HEADERS,
    OFFER_TYPE_OPTIONS,
    PAYMENT_TYPE_OPTIONS,
    VEHICLE_TYPE_OPTIONS,
)
from app.core.exceptions import ExcelGenerationError
from app.schemas.llm import VehicleIncentiveLLM


class ExcelService:
    def __init__(
        self,
        storage_dir: str | Path | None = None,
        timezone_name: str | None = None,
    ) -> None:
        self.storage_dir = Path(storage_dir or settings.local_storage_dir)
        self.timezone_name = timezone_name or settings.app_timezone

    @staticmethod
    def _slugify(value: str) -> str:
        value = value.strip().lower()
        value = re.sub(r"[^a-z0-9]+", "_", value)
        return value.strip("_") or "dealer"

    def build_file_name(self, dealer_name: str, created_at: datetime | None = None) -> str:
        local_now = created_at or datetime.now(ZoneInfo(self.timezone_name))
        dealer = self._slugify(dealer_name)
        # Example: norm_reeves_honda_irvine_august_11_tuesday.xlsx
        return f"{dealer}_{local_now.strftime('%B_%d_%A').lower()}.xlsx"

    @staticmethod
    def _value_for_excel(value):
        if value is None:
            return None
        if isinstance(value, Decimal):
            return float(value)
        if hasattr(value, "value"):
            return value.value
        return value

    def create_workbook(
        self,
        dealer_name: str,
        incentives: list[VehicleIncentiveLLM],
        source_url: str | None = None,
    ) -> tuple[str, str]:
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        file_name = self.build_file_name(dealer_name)
        final_path = self.storage_dir / file_name
        if final_path.exists():
            stem = final_path.stem
            suffix = final_path.suffix
            counter = 2
            while final_path.exists():
                file_name = f"{stem}_{counter}{suffix}"
                final_path = self.storage_dir / file_name
                counter += 1
        temp_path = self.storage_dir / f".{file_name}.tmp"

        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Monthly Vehicle Incentives"
            worksheet.sheet_view.showGridLines = False
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}1"

            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)

            for column_index, header in enumerate(EXCEL_HEADERS, start=1):
                cell = worksheet.cell(row=1, column=column_index, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            if source_url:
                worksheet["A1"].comment = Comment(
                    f"Source URL: {source_url}",
                    "Vehicle Offer Extraction API",
                )

            for row_index, incentive in enumerate(incentives[:5], start=2):
                for column_index, field_name in enumerate(EXCEL_FIELD_ORDER, start=1):
                    value = getattr(incentive, field_name)
                    cell = worksheet.cell(
                        row=row_index,
                        column=column_index,
                        value=self._value_for_excel(value),
                    )
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            currency_columns = [12, 13, 16, 17, 21, 22, 23]
            for column_index in currency_columns:
                for row in range(2, 7):
                    worksheet.cell(row=row, column=column_index).number_format = "$#,##0.00"

            for row in range(2, 7):
                worksheet.cell(row=row, column=18).number_format = "#,##0"
                worksheet.cell(row=row, column=19).number_format = "0.0%"
                finance_rate_cell = worksheet.cell(row=row, column=19)
                if finance_rate_cell.value is not None:
                    finance_rate_cell.value = finance_rate_cell.value / 100
                worksheet.cell(row=row, column=11).number_format = "@"

            offer_type_validation = DataValidation(
                type="list",
                formula1='"' + ",".join(OFFER_TYPE_OPTIONS) + '"',
                allow_blank=True,
            )
            vehicle_type_validation = DataValidation(
                type="list",
                formula1='"' + ",".join(VEHICLE_TYPE_OPTIONS) + '"',
                allow_blank=True,
            )
            payment_type_validation = DataValidation(
                type="list",
                formula1='"' + ",".join(PAYMENT_TYPE_OPTIONS) + '"',
                allow_blank=True,
            )

            worksheet.add_data_validation(offer_type_validation)
            worksheet.add_data_validation(vehicle_type_validation)
            worksheet.add_data_validation(payment_type_validation)
            offer_type_validation.add("B2:B6")
            vehicle_type_validation.add("D2:D6")
            payment_type_validation.add("O2:O6")

            # Basic review highlight for an impossible negative finance rate if a user edits it later.
            worksheet.conditional_formatting.add(
                "S2:S6",
                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                ),
            )

            widths = {
                1: 14,
                2: 28,
                3: 35,
                4: 30,
                5: 10,
                6: 15,
                7: 20,
                8: 20,
                9: 14,
                10: 18,
                11: 45,
                12: 16,
                13: 23,
                14: 20,
                15: 31,
                16: 18,
                17: 22,
                18: 18,
                19: 15,
                20: 20,
                21: 25,
                22: 18,
                23: 18,
                24: 70,
                25: 18,
                26: 30,
                27: 22,
            }
            for column_index, width in widths.items():
                worksheet.column_dimensions[get_column_letter(column_index)].width = width

            worksheet.row_dimensions[1].height = 40
            for row in range(2, 7):
                worksheet.row_dimensions[row].height = 55

            workbook.save(temp_path)
            os.replace(temp_path, final_path)
            return file_name, str(final_path)
        except Exception as exc:
            if temp_path.exists():
                temp_path.unlink(missing_ok=True)
            raise ExcelGenerationError(
                f"Failed to generate Excel workbook: {exc}"
            ) from exc
